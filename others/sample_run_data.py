import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

# Load the links from the raw_links.xlsx
wb = load_workbook("raw_links.xlsx")
ws = wb.active

# Create a new Excel workbook to store the scraped data
scraped_wb = Workbook()
scraped_ws = scraped_wb.active
scraped_ws.title = "Website Data"

# Write the header row for the Excel sheet
headers = ["Company Name", "Location", "City", "P.O Box", "Phone", "Mobile", "Website", "Products/Services"]
scraped_ws.append(headers)

# Iterate over each URL in the import_links column (assumed to be in column 'A')
for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):  # Start from row 2 to skip the header
    url = row[0]  # Extract the URL from the first column

    # Send an HTTP request to get the webpage content
    try:
        response = requests.get(url)
        soup = BeautifulSoup(response.content, "html.parser")

        # Find all the cards containing the required information
        cards = soup.find_all('div', class_='flex flex-grow flex-col w-full')  # Adjust as needed

        # Loop through each card and extract the relevant data
        for card in cards:
            # Extract company name (title)
            company_name = card.find('a', title=True).get_text(strip=True)

            # Extract location, city, and P.O. Box
            location = card.find('p', string="Location : ").find_next('span').get_text(strip=True) if card.find('p', string="Location : ") else "N/A"
            city = card.find('span', string="City : ").find_next('span').get_text(strip=True) if card.find('span', string="City : ") else "N/A"
            po_box = card.find('span', string="P.O Box : ").find_next('span').get_text(strip=True) if card.find('span', string="P.O Box : ") else "N/A"

            # Extract phone number from the href attribute
            phone = card.find('a', href=True, title="Click to call", id=lambda x: x and x.startswith('lblPhone'))
            phone_number = phone['href'].replace('tel:', '') if phone else "N/A"

            # Extract mobile number from the href attribute
            mobile = card.find('a', href=True, title="Click to call", id=lambda x: x and x.startswith('lblMobile'))
            mobile_number = mobile['href'].replace('tel:', '') if mobile else "N/A"

            # Extract website URL
            website = card.find('button', class_='listing_button')["data-url"] if card.find('button', class_='listing_button') else "N/A"

            # Extract products/services
            services = ", ".join([a.get_text(strip=True) for a in card.find_all('a', class_='font-bold')])

            # Write the extracted data to the Excel sheet
            row_data = [company_name, location, city, po_box, phone_number, mobile_number, website, services]
            scraped_ws.append(row_data)

    except Exception as e:
        print(f"Error scraping {url}: {e}")
        continue

# Save the scraped data to a new Excel file
scraped_wb.save("company_data.xlsx")

print("Data scraping and saving to Excel completed.")
