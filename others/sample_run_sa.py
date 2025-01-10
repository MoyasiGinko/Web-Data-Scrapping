import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

# Load the raw_links.xlsx to get the list of URLs
raw_links_wb = load_workbook("raw_links.xlsx")
raw_links_ws = raw_links_wb.active

# Create a new Excel workbook to store the scraped data
scraped_wb = Workbook()
scraped_ws = scraped_wb.active
scraped_ws.title = "Search Results Data"

# Write the header row for the Excel sheet
headers = ["Company Name", "Company Link", "Street Address", "Region", "Phone Number", "Product Link", "WhatsApp Chat", "Rating"]
scraped_ws.append(headers)

# Function to extract company data
def scrape_search_results(url):
    try:
        response = requests.get(url)
        response.raise_for_status()  # Check if request was successful
        soup = BeautifulSoup(response.content, "html.parser")

        # Find all search result cards
        result_cards = soup.find_all("ul", class_="f16 searchpageright pl-sm-0 pl-4")

        # Iterate over each card and extract information
        for card in result_cards:
            # Extract company name and link
            company_tag = card.find("a", style="color:#009900;")
            company_name = company_tag.get_text(strip=True)
            company_link = company_tag["href"]

            # Extract street address
            street_address_tag = card.find("li", itemprop="streetAddress")
            street_address = street_address_tag.get_text(strip=True) if street_address_tag else "N/A"

            # Extract region
            region_tag = card.find("li", itemprop="addressRegion")
            region = region_tag.get_text(strip=True) if region_tag else "N/A"

            # Extract phone number from "tel:" link
            phone_tag = card.find("li", itemprop="telephone").find("a", href=True)
            if phone_tag:
                phone_number = phone_tag["href"].replace("tel:", "").strip()  # Clean up the phone number
            else:
                phone_number = "N/A"

            # Extract product link
            product_link_tag = card.find("a", href=True, text="View our eShowRoom")
            product_link = product_link_tag["href"] if product_link_tag else "N/A"

            # Check if WhatsApp chat option is available
            whatsapp_tag = card.find("a", class_="openwhatsappenq")
            whatsapp_chat = "Yes" if whatsapp_tag else "No"

            # Extract rating
            rating_tag = card.find("div", class_="rateit")
            rating = rating_tag["data-rateit-value"] if rating_tag else "N/A"

            # Append data to the Excel sheet
            scraped_ws.append([company_name, company_link, street_address, region, phone_number, product_link, whatsapp_chat, rating])

    except Exception as e:
        print(f"Error scraping {url}: {e}")

# Function to get the URLs from the raw_links.xlsx file
def get_urls_from_excel():
    urls = []
    # Assuming 'import-LINKS' column contains the URLs
    for row in raw_links_ws.iter_rows(min_row=2, max_col=1, values_only=True):
        url = row[0]
        if url:
            urls.append(url)
    return urls

# Get the list of URLs from the raw_links.xlsx
urls_to_scrape = get_urls_from_excel()

# Iterate over each URL in the list and scrape data
for url in urls_to_scrape:
    scrape_search_results(url)

# Save the scraped data to a new Excel file
scraped_wb.save("search_results_data.xlsx")

print("Data scraping and saving to Excel completed.")
